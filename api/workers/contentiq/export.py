"""
ContentIQ Excel Export
======================
export_audit_excel(audit, pages) -> bytes
Produces a .xlsx with 3 sheets: Summary, All Pages, Briefs.
"""
from __future__ import annotations

import io
from typing import List

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter


# Verdict fill colours
_FILLS = {
    "KEEP":        PatternFill(fill_type="solid", fgColor="A9F3D1"),
    "UPDATE":      PatternFill(fill_type="solid", fgColor="FEF9C3"),
    "CONSOLIDATE": PatternFill(fill_type="solid", fgColor="FED7AA"),
    "DELETE":      PatternFill(fill_type="solid", fgColor="FEE2E2"),
}

_HEADER_FONT   = Font(bold=True)
_BOLD_FONT     = Font(bold=True)


def _auto_width(ws, min_w: int = 12, max_w: int = 60) -> None:
    """Set column widths based on cell content."""
    for col_cells in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            try:
                val = str(cell.value) if cell.value is not None else ""
                if len(val) > max_len:
                    max_len = len(val)
            except Exception:
                pass
        width = max(min_w, min(max_len + 2, max_w))
        ws.column_dimensions[col_letter].width = width


def _write_header(ws, headers: list) -> None:
    ws.append(headers)
    for cell in ws[ws.max_row]:
        cell.font = _HEADER_FONT
    ws.freeze_panes = ws["A2"]


def _pct(val) -> str:
    if val is None:
        return ""
    return f"{round(val * 100, 2)}%"


def _avg(values: list) -> float:
    vals = [v for v in values if v is not None]
    return round(sum(vals) / len(vals), 1) if vals else 0.0


def export_audit_excel(audit, pages: List) -> bytes:
    """
    Build and return the raw bytes of a ContentIQ audit Excel file.

    Parameters
    ----------
    audit : CiqAudit ORM instance
    pages : list[CiqPage] ORM instances
    """
    wb = Workbook()

    # ------------------------------------------------------------------ #
    # Sheet 1: Summary
    # ------------------------------------------------------------------ #
    ws_summary = wb.active
    ws_summary.title = "Summary"

    verdict_counts = {"KEEP": 0, "UPDATE": 0, "CONSOLIDATE": 0, "DELETE": 0}
    for p in pages:
        v = (p.verdict or "").upper()
        if v in verdict_counts:
            verdict_counts[v] += 1

    competitor_gap_count = sum(1 for p in pages if p.competitor_gap)

    avg_total     = _avg([p.score_total      for p in pages])
    avg_freshness = _avg([p.score_freshness  for p in pages])
    avg_geo       = _avg([p.score_geo        for p in pages])
    avg_eeat      = _avg([p.score_eeat       for p in pages])
    avg_seo       = _avg([p.score_seo_health for p in pages])

    summary_rows = [
        ("Audit Label",           audit.label),
        ("Domain",                audit.domain),
        ("Status",                audit.status),
        ("Created At",            audit.created_at.isoformat() if audit.created_at else ""),
        ("Finished At",           audit.finished_at.isoformat() if audit.finished_at else ""),
        ("Total URLs",            audit.total_urls),
        ("Scored URLs",           audit.scored_urls),
        ("KEEP Count",            verdict_counts["KEEP"]),
        ("UPDATE Count",          verdict_counts["UPDATE"]),
        ("CONSOLIDATE Count",     verdict_counts["CONSOLIDATE"]),
        ("DELETE Count",          verdict_counts["DELETE"]),
        ("Avg Total Score",       avg_total),
        ("Avg Freshness",         avg_freshness),
        ("Avg GEO",               avg_geo),
        ("Avg E-E-A-T",           avg_eeat),
        ("Avg SEO Health",        avg_seo),
        ("Competitor Gap Pages",  competitor_gap_count),
    ]

    # Header row
    ws_summary.append(["Field", "Value"])
    for cell in ws_summary[1]:
        cell.font = _HEADER_FONT
    ws_summary.freeze_panes = "A2"

    for label, value in summary_rows:
        ws_summary.append([label, value])

    _auto_width(ws_summary)

    # ------------------------------------------------------------------ #
    # Sheet 2: All Pages
    # ------------------------------------------------------------------ #
    ws_pages = wb.create_sheet("All Pages")

    page_headers = [
        "URL", "Title", "Word Count", "Status Code",
        "Verdict", "Score Total", "Score Freshness", "Score GEO",
        "Score E-E-A-T", "Score SEO Health",
        "GSC Clicks", "GSC Impressions", "GSC CTR", "GSC Position",
        "Ahrefs Traffic", "Ahrefs Keywords", "Ahrefs Backlinks", "Ahrefs DR",
        "Last Modified", "Competitor Gap", "Brief Generated",
    ]
    _write_header(ws_pages, page_headers)

    # Verdict column index (1-based) — "Verdict" is column 5
    VERDICT_COL = 5

    def _page_row(p):
        return [
            p.url,
            p.title or "",
            p.word_count,
            p.status_code,
            p.verdict or "",
            p.score_total,
            p.score_freshness,
            p.score_geo,
            p.score_eeat,
            p.score_seo_health,
            p.gsc_clicks,
            p.gsc_impressions,
            _pct(p.gsc_ctr),
            p.gsc_position,
            p.ahrefs_traffic,
            p.ahrefs_keywords,
            p.ahrefs_backlinks,
            p.ahrefs_dr,
            p.last_modified or "",
            "Yes" if p.competitor_gap else "No",
            "Yes" if p.brief_generated else "No",
        ]

    for p in pages:
        ws_pages.append(_page_row(p))
        row_num = ws_pages.max_row
        verdict = (p.verdict or "").upper()
        fill = _FILLS.get(verdict)
        if fill:
            ws_pages.cell(row=row_num, column=VERDICT_COL).fill = fill

    # Competitor gap section
    gap_pages = [p for p in pages if p.competitor_gap]
    if gap_pages:
        # Blank row
        ws_pages.append([""] * len(page_headers))

        # Bold header
        ws_pages.append(["COMPETITOR GAP PAGES"] + [""] * (len(page_headers) - 1))
        header_row_num = ws_pages.max_row
        ws_pages.cell(row=header_row_num, column=1).font = _BOLD_FONT

        for p in gap_pages:
            ws_pages.append(_page_row(p))
            row_num = ws_pages.max_row
            verdict = (p.verdict or "").upper()
            fill = _FILLS.get(verdict)
            if fill:
                ws_pages.cell(row=row_num, column=VERDICT_COL).fill = fill

    _auto_width(ws_pages)

    # ------------------------------------------------------------------ #
    # Sheet 3: Briefs
    # ------------------------------------------------------------------ #
    ws_briefs = wb.create_sheet("Briefs")

    brief_headers = ["URL", "Verdict", "Score Total", "Brief Content"]
    _write_header(ws_briefs, brief_headers)

    for p in pages:
        if p.brief_generated:
            ws_briefs.append([
                p.url,
                p.verdict or "",
                p.score_total,
                p.brief_content or "",
            ])
            row_num = ws_briefs.max_row
            verdict = (p.verdict or "").upper()
            fill = _FILLS.get(verdict)
            if fill:
                ws_briefs.cell(row=row_num, column=2).fill = fill

    # Wrap brief content column
    brief_col_letter = get_column_letter(4)
    for row in ws_briefs.iter_rows(min_row=2, min_col=4, max_col=4):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    _auto_width(ws_briefs)

    # ------------------------------------------------------------------ #
    # Serialise to bytes
    # ------------------------------------------------------------------ #
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
